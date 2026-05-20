"""
Excel操作节点

支持Excel文件的读取、写入和创建操作
使用openpyxl库处理.xlsx文件
"""

import os
from typing import Any, Dict

from ..models.schemas import NodeDefinition, NodeInput, NodeOutput, NodeType, InputType
from ..models.context import ExecutionContext
from .base import BaseNode


def _ensure_openpyxl():
    """确保openpyxl库可用"""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "需要安装openpyxl库: pip install openpyxl\n"
            "该库用于处理Excel(.xlsx)文件"
        )


class ExcelReadNode(BaseNode):
    """
    读取Excel节点

    读取Excel文件中的数据，支持指定工作表和范围。

    参数说明:
    - file_path: Excel文件路径（.xlsx格式）
    - sheet_name: 工作表名称（默认第一个）
    - range: 读取范围（可选，如A1:C10）
    - has_header: 是否有表头行（默认true）
    """

    def _create_definition(self) -> NodeDefinition:
        return NodeDefinition(
            type=NodeType.EXCEL_READ,
            name="读取Excel",
            description="读取Excel文件中的数据，返回行列表",
            category="Excel",
            inputs=[
                NodeInput(
                    name="file_path",
                    label="文件路径",
                    type=InputType.FILE_PATH,
                    required=True,
                    description="Excel文件路径（.xlsx格式）"
                ),
                NodeInput(
                    name="sheet_name",
                    label="工作表名称",
                    type=InputType.TEXT,
                    required=False,
                    default="",
                    description="工作表名称，留空则读取第一个工作表"
                ),
                NodeInput(
                    name="range",
                    label="读取范围",
                    type=InputType.TEXT,
                    required=False,
                    default="",
                    description="读取范围，如: A1:C10，留空则读取全部数据"
                ),
                NodeInput(
                    name="has_header",
                    label="包含表头",
                    type=InputType.BOOLEAN,
                    required=False,
                    default=True,
                    description="第一行是否为表头"
                ),
            ],
            outputs=[
                NodeOutput(
                    key="excel_data",
                    label="Excel数据",
                    description="读取结果字典，包含headers(表头)、rows(数据行)、row_count(行数)"
                ),
            ]
        )

    def execute(self, inputs: Dict[str, Any], context: ExecutionContext) -> Dict[str, Any]:
        openpyxl = _ensure_openpyxl()

        file_path = self.get_required_input(inputs, "file_path")
        sheet_name = self.get_input_value(inputs, "sheet_name", "")
        read_range = self.get_input_value(inputs, "range", "")
        has_header = self.get_input_value(inputs, "has_header", True)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # 获取工作表
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在，可用工作表: {wb.sheetnames}")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # 读取数据
            if read_range:
                cells = ws[read_range]
            else:
                cells = ws.iter_rows()

            headers = []
            rows = []

            for i, row in enumerate(cells):
                row_data = [cell.value for cell in row]
                if i == 0 and has_header:
                    headers = row_data
                else:
                    if has_header and headers:
                        # 返回字典列表
                        rows.append(dict(zip(headers, row_data)))
                    else:
                        # 返回列表的列表
                        rows.append(row_data)

            wb.close()

            return {
                "excel_data": {
                    "success": True,
                    "headers": headers,
                    "rows": rows,
                    "row_count": len(rows),
                    "sheet_name": ws.title,
                    "file_path": file_path
                }
            }
        except Exception as e:
            return {
                "excel_data": {
                    "success": False,
                    "error": str(e),
                    "rows": [],
                    "row_count": 0
                }
            }


class ExcelWriteNode(BaseNode):
    """
    写入Excel节点

    向现有Excel文件写入数据。

    参数说明:
    - file_path: Excel文件路径
    - sheet_name: 工作表名称（默认第一个）
    - data: 要写入的数据（JSON格式的列表）
    - start_cell: 起始单元格（默认A1）
    - append: 是否追加模式（默认false，覆盖写入）
    """

    def _create_definition(self) -> NodeDefinition:
        return NodeDefinition(
            type=NodeType.EXCEL_WRITE,
            name="写入Excel",
            description="向Excel文件写入数据",
            category="Excel",
            inputs=[
                NodeInput(
                    name="file_path",
                    label="文件路径",
                    type=InputType.FILE_PATH,
                    required=True,
                    description="Excel文件路径（.xlsx格式）"
                ),
                NodeInput(
                    name="sheet_name",
                    label="工作表名称",
                    type=InputType.TEXT,
                    required=False,
                    default="",
                    description="工作表名称，留空则使用第一个工作表"
                ),
                NodeInput(
                    name="data",
                    label="写入数据",
                    type=InputType.CODE,
                    required=True,
                    description="要写入的数据（JSON格式），如: [['姓名','年龄'],['张三',25],['李四',30]]"
                ),
                NodeInput(
                    name="start_cell",
                    label="起始单元格",
                    type=InputType.TEXT,
                    required=False,
                    default="A1",
                    description="写入起始位置，如: A1, B3"
                ),
                NodeInput(
                    name="append",
                    label="追加模式",
                    type=InputType.BOOLEAN,
                    required=False,
                    default=False,
                    description="是否在现有数据后追加（false=覆盖写入）"
                ),
            ],
            outputs=[
                NodeOutput(
                    key="write_result",
                    label="写入结果",
                    description="写入结果信息"
                ),
            ]
        )

    def execute(self, inputs: Dict[str, Any], context: ExecutionContext) -> Dict[str, Any]:
        import json
        openpyxl = _ensure_openpyxl()

        file_path = self.get_required_input(inputs, "file_path")
        sheet_name = self.get_input_value(inputs, "sheet_name", "")
        data_str = self.get_required_input(inputs, "data")
        start_cell = self.get_input_value(inputs, "start_cell", "A1")
        append = self.get_input_value(inputs, "append", False)

        # 解析数据
        try:
            data = json.loads(data_str) if isinstance(data_str, str) else data_str
        except json.JSONDecodeError as e:
            raise ValueError(f"数据格式错误，需要JSON格式: {e}")

        if not isinstance(data, list):
            raise ValueError("数据必须是列表格式")

        try:
            # 打开或创建工作簿
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
            else:
                wb = openpyxl.Workbook()

            # 获取工作表
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
            else:
                ws = wb.active

            # 确定起始位置
            from openpyxl.utils import coordinate_to_tuple
            try:
                start_row, start_col = coordinate_to_tuple(start_cell)
            except Exception:
                start_row, start_col = 1, 1

            # 追加模式：找到最后一行
            if append:
                start_row = ws.max_row + 1

            # 写入数据
            rows_written = 0
            for i, row_data in enumerate(data):
                if isinstance(row_data, list):
                    for j, value in enumerate(row_data):
                        ws.cell(row=start_row + i, column=start_col + j, value=value)
                elif isinstance(row_data, dict):
                    for j, (key, value) in enumerate(row_data.items()):
                        ws.cell(row=start_row + i, column=start_col + j, value=value)
                rows_written += 1

            wb.save(file_path)
            wb.close()

            return {
                "write_result": {
                    "success": True,
                    "file_path": file_path,
                    "rows_written": rows_written,
                    "message": f"成功写入 {rows_written} 行数据到 {file_path}"
                }
            }
        except Exception as e:
            return {
                "write_result": {
                    "success": False,
                    "error": str(e)
                }
            }


class ExcelCreateNode(BaseNode):
    """
    创建Excel节点

    创建新的Excel文件。

    参数说明:
    - file_path: 新文件路径
    - sheet_names: 工作表名称列表（JSON格式，默认["Sheet1"]）
    - data: 初始数据（可选，JSON格式，写入第一个工作表）
    - headers: 表头行（可选，JSON格式）
    """

    def _create_definition(self) -> NodeDefinition:
        return NodeDefinition(
            type=NodeType.EXCEL_CREATE,
            name="创建Excel",
            description="创建新的Excel文件",
            category="Excel",
            inputs=[
                NodeInput(
                    name="file_path",
                    label="文件路径",
                    type=InputType.FILE_PATH,
                    required=True,
                    description="新Excel文件路径（.xlsx格式）"
                ),
                NodeInput(
                    name="sheet_names",
                    label="工作表名称",
                    type=InputType.TEXT,
                    required=False,
                    default='["Sheet1"]',
                    description="工作表名称列表（JSON格式），如: ['Sheet1','数据','汇总']"
                ),
                NodeInput(
                    name="headers",
                    label="表头",
                    type=InputType.TEXT,
                    required=False,
                    default="[]",
                    description="表头行（JSON格式），如: ['姓名','年龄','部门']"
                ),
                NodeInput(
                    name="data",
                    label="初始数据",
                    type=InputType.CODE,
                    required=False,
                    default="[]",
                    description="初始数据（JSON格式），如: [['张三',25,'技术部'],['李四',30,'市场部']]"
                ),
            ],
            outputs=[
                NodeOutput(
                    key="create_result",
                    label="创建结果",
                    description="创建结果信息"
                ),
            ]
        )

    def execute(self, inputs: Dict[str, Any], context: ExecutionContext) -> Dict[str, Any]:
        import json
        openpyxl = _ensure_openpyxl()

        file_path = self.get_required_input(inputs, "file_path")
        sheet_names_str = self.get_input_value(inputs, "sheet_names", '["Sheet1"]')
        headers_str = self.get_input_value(inputs, "headers", "[]")
        data_str = self.get_input_value(inputs, "data", "[]")

        # 解析参数
        try:
            sheet_names = json.loads(sheet_names_str) if isinstance(sheet_names_str, str) else sheet_names_str
            headers = json.loads(headers_str) if isinstance(headers_str, str) else headers_str
            data = json.loads(data_str) if isinstance(data_str, str) else data_str
        except json.JSONDecodeError as e:
            raise ValueError(f"参数格式错误，需要JSON格式: {e}")

        try:
            wb = openpyxl.Workbook()

            # 创建工作表
            for i, name in enumerate(sheet_names):
                if i == 0:
                    ws = wb.active
                    ws.title = name
                else:
                    ws = wb.create_sheet(name)

            # 写入第一个工作表
            ws = wb[sheet_names[0]]
            current_row = 1

            # 写入表头
            if headers:
                for j, header in enumerate(headers):
                    ws.cell(row=current_row, column=j + 1, value=header)
                current_row += 1

            # 写入数据
            for row_data in data:
                if isinstance(row_data, list):
                    for j, value in enumerate(row_data):
                        ws.cell(row=current_row, column=j + 1, value=value)
                current_row += 1

            # 创建目录
            os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)

            wb.save(file_path)
            wb.close()

            return {
                "create_result": {
                    "success": True,
                    "file_path": file_path,
                    "sheet_names": sheet_names,
                    "message": f"成功创建Excel文件: {file_path}"
                }
            }
        except Exception as e:
            return {
                "create_result": {
                    "success": False,
                    "error": str(e)
                }
            }
