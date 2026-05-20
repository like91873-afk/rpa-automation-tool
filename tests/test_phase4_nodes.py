"""
Phase 4 测试 - 数据库、Excel、Web、延迟节点
"""

import json
import os
import sqlite3
import tempfile
import time

import pytest

from rpa_engine.models.schemas import Flow, NodeInstance, Connection, NodeType
from rpa_engine.models.context import ExecutionContext
from rpa_engine.engine import ExecutionEngine
from rpa_engine.nodes import NODE_REGISTRY, get_node_class, get_all_node_definitions
from rpa_engine.nodes.database_nodes import DBConnectNode, DBQueryNode, DBExecuteNode
from rpa_engine.nodes.excel_nodes import ExcelReadNode, ExcelWriteNode, ExcelCreateNode
from rpa_engine.nodes.web_nodes import HTTPRequestNode, WebScrapeNode
from rpa_engine.nodes.delay_nodes import DelayNode, WaitForNode


# ==================== 数据库节点测试 ====================

class TestDBConnectNode:
    """数据库连接节点测试"""

    def test_definition(self):
        node = DBConnectNode()
        defn = node.definition
        assert defn.type == NodeType.DB_CONNECT
        assert defn.name == "数据库连接"
        assert defn.category == "数据库"
        assert len(defn.inputs) >= 1
        assert len(defn.outputs) >= 1

    def test_connect_sqlite(self):
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
            db_path = f.name

        try:
            ctx = ExecutionContext()
            node = DBConnectNode()
            result = node.execute({"db_path": db_path}, ctx)

            assert result["db_connect_result"]["success"] is True
            assert result["db_connection"] is not None
            assert db_path in result["db_connect_result"]["db_path"]

            # 关闭连接
            result["db_connection"].close()
        finally:
            os.unlink(db_path)

    def test_connect_in_flow(self):
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
            db_path = f.name

        try:
            flow = Flow(
                id="test-db-connect",
                name="测试数据库连接",
                nodes=[
                    NodeInstance(id="n1", type="db_connect", inputs={"db_path": db_path}),
                ],
                connections=[]
            )
            engine = ExecutionEngine()
            result = engine.execute_flow(flow)
            assert result.status.value == "completed"
            assert result.variables.get("db_connect_result", {}).get("success") is True
        finally:
            os.unlink(db_path)


class TestDBQueryNode:
    """数据库查询节点测试"""

    def test_definition(self):
        node = DBQueryNode()
        defn = node.definition
        assert defn.type == NodeType.DB_QUERY
        assert defn.name == "数据库查询"

    def test_query_execution(self):
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
            db_path = f.name

        try:
            # 创建测试数据
            conn = sqlite3.connect(db_path)
            conn.execute("CREATE TABLE users (id INTEGER PRIMARY KEY, name TEXT, age INTEGER)")
            conn.execute("INSERT INTO users VALUES (1, 'Alice', 30)")
            conn.execute("INSERT INTO users VALUES (2, 'Bob', 25)")
            conn.commit()
            conn.close()

            # 测试查询
            flow = Flow(
                id="test-db-query",
                name="测试数据库查询",
                nodes=[
                    NodeInstance(id="n1", type="db_connect", inputs={"db_path": db_path}),
                    NodeInstance(id="n2", type="db_query", inputs={
                        "db_connection": "${db_connection}",
                        "sql": "SELECT * FROM users ORDER BY id",
                        "params": "[]"
                    }),
                ],
                connections=[
                    Connection(id="c1", source_node_id="n1", target_node_id="n2")
                ]
            )
            engine = ExecutionEngine()
            result = engine.execute_flow(flow)

            assert result.status.value == "completed"
            query_result = result.variables.get("query_result", {})
            assert query_result.get("success") is True
            assert query_result.get("row_count") == 2
            assert query_result.get("rows")[0]["name"] == "Alice"
        finally:
            os.unlink(db_path)


class TestDBExecuteNode:
    """数据库执行节点测试"""

    def test_definition(self):
        node = DBExecuteNode()
        defn = node.definition
        assert defn.type == NodeType.DB_EXECUTE
        assert defn.name == "数据库执行"

    def test_insert_execution(self):
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
            db_path = f.name

        try:
            # 创建表
            conn = sqlite3.connect(db_path)
            conn.execute("CREATE TABLE items (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT)")
            conn.commit()
            conn.close()

            # 测试插入
            flow = Flow(
                id="test-db-insert",
                name="测试数据库插入",
                nodes=[
                    NodeInstance(id="n1", type="db_connect", inputs={"db_path": db_path}),
                    NodeInstance(id="n2", type="db_execute", inputs={
                        "db_connection": "${db_connection}",
                        "sql": "INSERT INTO items (name) VALUES (?)",
                        "params": '["TestItem"]'
                    }),
                ],
                connections=[
                    Connection(id="c1", source_node_id="n1", target_node_id="n2")
                ]
            )
            engine = ExecutionEngine()
            result = engine.execute_flow(flow)

            assert result.status.value == "completed"
            exec_result = result.variables.get("execute_result", {})
            assert exec_result.get("success") is True
            assert exec_result.get("affected_rows") == 1

            # 验证数据
            conn = sqlite3.connect(db_path)
            row = conn.execute("SELECT name FROM items WHERE id=1").fetchone()
            assert row[0] == "TestItem"
            conn.close()
        finally:
            os.unlink(db_path)


# ==================== Excel节点测试 ====================

class TestExcelCreateNode:
    """创建Excel节点测试"""

    def test_definition(self):
        node = ExcelCreateNode()
        defn = node.definition
        assert defn.type == NodeType.EXCEL_CREATE
        assert defn.name == "创建Excel"
        assert defn.category == "Excel"

    def test_create_excel(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            xlsx_path = f.name

        try:
            ctx = ExecutionContext()
            node = ExcelCreateNode()
            result = node.execute({
                "file_path": xlsx_path,
                "sheet_names": '["测试表"]',
                "headers": '["姓名", "年龄"]',
                "data": '[["张三", 25], ["李四", 30]]'
            }, ctx)

            assert result["create_result"]["success"] is True
            assert os.path.exists(xlsx_path)

            # 验证内容
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            assert ws.title == "测试表"
            assert ws.cell(1, 1).value == "姓名"
            assert ws.cell(2, 1).value == "张三"
            assert ws.cell(3, 2).value == 30
            wb.close()
        finally:
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_create_in_flow(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            xlsx_path = f.name

        try:
            flow = Flow(
                id="test-excel-create",
                name="测试创建Excel",
                nodes=[
                    NodeInstance(id="n1", type="excel_create", inputs={
                        "file_path": xlsx_path,
                        "headers": '["A", "B"]',
                        "data": '[[1, 2], [3, 4]]'
                    }),
                ],
                connections=[]
            )
            engine = ExecutionEngine()
            result = engine.execute_flow(flow)
            assert result.status.value == "completed"
        finally:
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)


class TestExcelReadWriteNode:
    """Excel读写节点测试"""

    def test_write_and_read(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            xlsx_path = f.name

        try:
            # 先创建
            ctx = ExecutionContext()
            create_node = ExcelCreateNode()
            create_node.execute({
                "file_path": xlsx_path,
                "headers": '["Name", "Value"]',
                "data": '[["A", 100], ["B", 200]]'
            }, ctx)

            # 读取
            read_node = ExcelReadNode()
            result = read_node.execute({
                "file_path": xlsx_path,
                "has_header": True
            }, ctx)

            assert result["excel_data"]["success"] is True
            assert result["excel_data"]["headers"] == ["Name", "Value"]
            assert len(result["excel_data"]["rows"]) == 2
            assert result["excel_data"]["rows"][0]["Name"] == "A"
            assert result["excel_data"]["rows"][1]["Value"] == 200
        finally:
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_write_append(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            xlsx_path = f.name

        try:
            ctx = ExecutionContext()
            # 创建
            create_node = ExcelCreateNode()
            create_node.execute({
                "file_path": xlsx_path,
                "data": '[["初始数据"]]'
            }, ctx)

            # 追加写入
            write_node = ExcelWriteNode()
            result = write_node.execute({
                "file_path": xlsx_path,
                "data": '[["追加数据1"], ["追加数据2"]]',
                "append": True
            }, ctx)

            assert result["write_result"]["success"] is True

            # 验证
            read_node = ExcelReadNode()
            read_result = read_node.execute({
                "file_path": xlsx_path,
                "has_header": False
            }, ctx)
            assert read_result["excel_data"]["row_count"] == 3
        finally:
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)


# ==================== Web节点测试 ====================

class TestHTTPRequestNode:
    """HTTP请求节点测试"""

    def test_definition(self):
        node = HTTPRequestNode()
        defn = node.definition
        assert defn.type == NodeType.HTTP_REQUEST
        assert defn.name == "HTTP请求"
        assert defn.category == "Web"

    def test_get_request(self):
        ctx = ExecutionContext()
        node = HTTPRequestNode()
        # 使用一个公开的测试API
        result = node.execute({
            "url": "https://httpbin.org/get",
            "method": "GET",
            "headers": "{}",
            "timeout": 10
        }, ctx)

        assert result["response"]["success"] is True
        assert result["response"]["status_code"] == 200
        assert result["response"]["json"] is not None

    def test_invalid_url(self):
        ctx = ExecutionContext()
        node = HTTPRequestNode()
        result = node.execute({
            "url": "http://localhost:1",  # 不可达
            "method": "GET",
            "timeout": 2
        }, ctx)

        assert result["response"]["success"] is False


class TestWebScrapeNode:
    """网页抓取节点测试"""

    def test_definition(self):
        node = WebScrapeNode()
        defn = node.definition
        assert defn.type == NodeType.WEB_SCRAPE
        assert defn.name == "网页抓取"

    def test_scrape_text(self):
        ctx = ExecutionContext()
        node = WebScrapeNode()
        result = node.execute({
            "url": "https://httpbin.org/html",
            "extract": "text",
            "timeout": 10
        }, ctx)

        assert result["scrape_result"]["success"] is True
        assert len(result["scrape_result"]["content"]) > 0


# ==================== 延迟节点测试 ====================

class TestDelayNode:
    """延迟节点测试"""

    def test_definition(self):
        node = DelayNode()
        defn = node.definition
        assert defn.type == NodeType.DELAY
        assert defn.name == "延迟执行"
        assert defn.category == "控制流"

    def test_short_delay(self):
        ctx = ExecutionContext()
        node = DelayNode()
        start = time.time()
        result = node.execute({"seconds": 0.1, "message": "测试延迟"}, ctx)
        elapsed = time.time() - start

        assert result["delay_result"]["success"] is True
        assert result["delay_result"]["requested_seconds"] == 0.1
        assert elapsed >= 0.1

    def test_delay_in_flow(self):
        flow = Flow(
            id="test-delay",
            name="测试延迟",
            nodes=[
                NodeInstance(id="n1", type="delay", inputs={"seconds": 0.1}),
            ],
            connections=[]
        )
        engine = ExecutionEngine()
        result = engine.execute_flow(flow)
        assert result.status.value == "completed"

    def test_negative_delay_error(self):
        ctx = ExecutionContext()
        node = DelayNode()
        with pytest.raises(ValueError, match="不能为负数"):
            node.execute({"seconds": -1}, ctx)


class TestWaitForNode:
    """等待条件节点测试"""

    def test_definition(self):
        node = WaitForNode()
        defn = node.definition
        assert defn.type == NodeType.WAIT_FOR
        assert defn.name == "等待条件"

    def test_wait_variable_exists(self):
        ctx = ExecutionContext()
        ctx.set_variable("target_var", "hello")

        node = WaitForNode()
        result = node.execute({
            "condition_type": "variable_exists",
            "variable_name": "target_var",
            "timeout": 5,
            "poll_interval": 0.1
        }, ctx)

        assert result["wait_result"]["success"] is True
        assert result["wait_result"]["timeout"] is False

    def test_wait_variable_true(self):
        ctx = ExecutionContext()
        ctx.set_variable("flag", True)

        node = WaitForNode()
        result = node.execute({
            "condition_type": "variable_true",
            "variable_name": "flag",
            "timeout": 5,
            "poll_interval": 0.1
        }, ctx)

        assert result["wait_result"]["success"] is True

    def test_wait_expression(self):
        ctx = ExecutionContext()
        ctx.set_variable("count", 10)

        node = WaitForNode()
        result = node.execute({
            "condition_type": "python_expression",
            "expression": "count > 5",
            "timeout": 5,
            "poll_interval": 0.1
        }, ctx)

        assert result["wait_result"]["success"] is True

    def test_wait_timeout(self):
        ctx = ExecutionContext()

        node = WaitForNode()
        result = node.execute({
            "condition_type": "variable_exists",
            "variable_name": "never_exists",
            "timeout": 0.3,
            "poll_interval": 0.1
        }, ctx)

        assert result["wait_result"]["success"] is False
        assert result["wait_result"]["timeout"] is True


# ==================== 集成测试 ====================

class TestPhase4Integration:
    """Phase 4 集成测试"""

    def test_node_registry_completeness(self):
        """测试节点注册表完整性"""
        expected_nodes = [
            # Phase 1
            "python_exec", "python_script",
            "file_open", "file_read", "file_write", "directory_list",
            "system_cmd", "powershell", "computer_info",
            # Phase 2
            "condition", "loop", "math_operation", "string_operation",
            # Phase 3
            "sftp_connect", "sftp_upload", "sftp_download", "sftp_new_file", "sftp_write_file",
            "ftp_connect", "ftp_list_dir",
            "xml_save", "path_exists",
            # Phase 4
            "db_connect", "db_query", "db_execute",
            "excel_read", "excel_write", "excel_create",
            "http_request", "web_scrape",
            "delay", "wait_for",
        ]
        definitions = get_all_node_definitions()
        assert len(definitions) >= len(expected_nodes)

        registered_types = [d.type.value for d in definitions]
        for node_type in expected_nodes:
            assert node_type in registered_types, f"节点类型未注册: {node_type}"

    def test_node_type_enum_completeness(self):
        """测试NodeType枚举完整性"""
        expected_types = [
            "db_connect", "db_query", "db_execute",
            "excel_read", "excel_write", "excel_create",
            "http_request", "web_scrape",
            "delay", "wait_for",
        ]
        for type_name in expected_types:
            assert hasattr(NodeType, type_name.upper()), f"NodeType枚举缺失: {type_name}"

    def test_total_node_count(self):
        """测试总节点数量"""
        definitions = get_all_node_definitions()
        assert len(definitions) == 32  # 22 (Phase 1-3) + 10 (Phase 4)

    def test_database_flow(self):
        """测试数据库完整流程"""
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
            db_path = f.name

        try:
            flow = Flow(
                id="test-db-flow",
                name="数据库完整流程",
                nodes=[
                    NodeInstance(id="n1", type="db_connect", inputs={"db_path": db_path}),
                    NodeInstance(id="n2", type="db_execute", inputs={
                        "db_connection": "${db_connection}",
                        "sql": "CREATE TABLE test (id INTEGER PRIMARY KEY, value TEXT)"
                    }),
                    NodeInstance(id="n3", type="db_execute", inputs={
                        "db_connection": "${db_connection}",
                        "sql": "INSERT INTO test (value) VALUES (?)",
                        "params": '["hello"]'
                    }),
                    NodeInstance(id="n4", type="db_query", inputs={
                        "db_connection": "${db_connection}",
                        "sql": "SELECT * FROM test"
                    }),
                ],
                connections=[
                    Connection(id="c1", source_node_id="n1", target_node_id="n2"),
                    Connection(id="c2", source_node_id="n2", target_node_id="n3"),
                    Connection(id="c3", source_node_id="n3", target_node_id="n4"),
                ]
            )

            engine = ExecutionEngine()
            result = engine.execute_flow(flow)

            assert result.status.value == "completed"
            query_result = result.variables.get("query_result", {})
            assert query_result.get("row_count") == 1
            assert query_result.get("rows")[0]["value"] == "hello"
        finally:
            os.unlink(db_path)

    def test_excel_flow(self):
        """测试Excel完整流程"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            xlsx_path = f.name

        try:
            flow = Flow(
                id="test-excel-flow",
                name="Excel完整流程",
                nodes=[
                    NodeInstance(id="n1", type="excel_create", inputs={
                        "file_path": xlsx_path,
                        "headers": '["产品", "价格"]',
                        "data": '[["产品A", 100], ["产品B", 200]]'
                    }),
                    NodeInstance(id="n2", type="excel_read", inputs={
                        "file_path": xlsx_path,
                        "has_header": True
                    }),
                ],
                connections=[
                    Connection(id="c1", source_node_id="n1", target_node_id="n2"),
                ]
            )

            engine = ExecutionEngine()
            result = engine.execute_flow(flow)

            assert result.status.value == "completed"
            excel_data = result.variables.get("excel_data", {})
            assert excel_data.get("success") is True
            assert len(excel_data.get("rows", [])) == 2
        finally:
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_delay_flow(self):
        """测试延迟流程"""
        flow = Flow(
            id="test-delay-flow",
            name="延迟流程",
            nodes=[
                NodeInstance(id="n1", type="python_exec", inputs={
                    "python_code": "x = 42"
                }),
                NodeInstance(id="n2", type="delay", inputs={"seconds": 0.05}),
                NodeInstance(id="n3", type="python_exec", inputs={
                    "python_code": "result = x * 2"
                }),
            ],
            connections=[
                Connection(id="c1", source_node_id="n1", target_node_id="n2"),
                Connection(id="c2", source_node_id="n2", target_node_id="n3"),
            ]
        )

        engine = ExecutionEngine()
        result = engine.execute_flow(flow)

        assert result.status.value == "completed"
        assert result.variables.get("result") == 84
